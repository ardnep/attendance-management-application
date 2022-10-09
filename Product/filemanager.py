import os
import sys
import subprocess
import openpyxl
import pickle
import ntpath
from datetime import datetime
from plyer import filechooser
from users import User


class FileManager:
    @staticmethod
    def open_file():
        try:
            return filechooser.open_file(title="Pick a file...", filters=[("*.txt")])[0]
        except TypeError:
            return 'None'

    @staticmethod
    def find_file(name_of_record, path='.', specific=False):
        result = []
        files_lower = []
        path = os.path.expanduser(path)
        if specific:
            for root, dirs, files in os.walk(path):
                files_lower[:] = [file.lower() for file in files]
                for index, file in enumerate(files_lower):
                    if f'{name_of_record.lower()}_ta_app.xlsx' == file:
                        result.append(os.path.join(root, files[index]))
        else:
            for root, dirs, files in os.walk(path):
                files_lower[:] = [file.lower() for file in files]
                for index, file in enumerate(files_lower):
                    if name_of_record.lower() in file and '_ta_app.xlsx' in file:
                        result.append(os.path.join(root, files[index]))
        return result

    @staticmethod
    def create_excel_file(names, name_of_record):
        names = FileManager.process_names(names)

        wb = openpyxl.Workbook()
        sheet = wb.active

        # add serial number
        (sheet.cell(row=1, column=1)).value = 'SN'
        for i in range(1, len(names) + 1):
            (sheet.cell(row=i + 1, column=1)).value = i

        # add names
        (sheet.cell(row=1, column=2)).value = 'Name'
        for i in names:
            (sheet.cell(row=(names.index(i) + 2), column=2)).value = i

        # save file
        excel_file_path = os.path.expanduser(f'users/{User.current_user}/{name_of_record}_ta_app.xlsx')
        wb.save(excel_file_path)
        FileManager.start_file(excel_file_path)

    @staticmethod
    def update_excel_file(chat_file, excel_file):
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        take_attendance = AttendanceHandler(sheet)
        present_students = take_attendance.get_present_students(chat_file)

        # making new column and adding current date
        new_column = sheet.max_column + 1
        current_date = str(datetime.date(datetime.now()))
        (sheet.cell(row=1, column=new_column)).value = current_date

        present_students[:] = [student.lower() for student in present_students]

        # adding presence
        for i in range(2, sheet.max_row + 1):  # plus one is done as it needs to go up to the last value of sheet.max_row
            if sheet.cell(row=i, column=2).value.lower() in present_students:
                (sheet.cell(row=i, column=new_column)).value = "P"

        wb.save(excel_file)
        FileManager.start_file(excel_file)

    @staticmethod
    def add_students_excel(names_of_students, excel_file):
        additional_names = FileManager.process_names(names_of_students)

        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        handler = AttendanceHandler(sheet)
        handler.get_names_excel()

        current_names = handler.excel_names

        updated_record = {}

        max_column = sheet.max_column
        max_row = sheet.max_row

        for name in additional_names:
            updated_record[name] = [None]*(max_column - 2)

        current_row = 2
        while current_row <= max_row:
            name_of_student = (sheet.cell(row=current_row, column=2)).value
            current_column = 3
            presence = []
            while current_column <= max_column:
                presence.append((sheet.cell(row=current_row, column=current_column)).value)
                current_column += 1

            updated_record[name_of_student] = presence
            current_row += 1

        updated_names = [None]*(len(additional_names) + len(current_names))

        # merging the two sorted lists

        i = 0
        j = 0
        k = 0
        while i < len(additional_names) and j < len(current_names):
            if additional_names[i] < current_names[j]:
                updated_names[k] = additional_names[i]
                k += 1
                i += 1
            else:
                updated_names[k] = current_names[j]
                k += 1
                j += 1

        while i < len(additional_names):
            updated_names[k] = additional_names[i]
            k += 1
            i += 1

        while j < len(current_names):
            updated_names[k] = current_names[j]
            k += 1
            j += 1

        # Updated SN
        for sn in range(len(current_names), len(updated_names) + 1):
            (sheet.cell(row=sn + 1, column=1)).value = sn

        # Updated Names and Presence
        for name in updated_names:
            current_row = updated_names.index(name) + 2
            current_column = 2
            (sheet.cell(row=current_row, column=current_column)).value = name
            for presence in updated_record[name]:
                current_column += 1
                (sheet.cell(row=current_row, column=current_column)).value = presence

        wb.save(excel_file)
        FileManager.start_file(excel_file)

    @staticmethod
    def remove_students_excel(names_of_students, excel_file):
        names_to_remove = FileManager.process_names(names_of_students)

        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        handler = AttendanceHandler(sheet)
        handler.get_names_excel()

        current_names = handler.excel_names

        updated_record = {}

        max_column = sheet.max_column
        max_row = sheet.max_row

        current_row = 2
        while current_row <= max_row:
            name_of_student = (sheet.cell(row=current_row, column=2)).value
            current_column = 3
            presence = []
            while current_column <= max_column:
                presence.append((sheet.cell(row=current_row, column=current_column)).value)
                current_column += 1

            updated_record[name_of_student] = presence
            current_row += 1

        updated_names = []
        for name in current_names:
            print(name)
            if name not in names_to_remove:
                updated_names.append(name)

        print(updated_record)
        print(updated_names)

        # Clearing Sheet
        for i in range(2, max_row + 1):
            for j in range(1, max_column + 1):
                (sheet.cell(row=i, column=j)).value = None

        # Updated SN
        for sn in range(1, len(updated_names) + 1):
            (sheet.cell(row=sn + 1, column=1)).value = sn

        # Updated Names and Presence
        for name in updated_names:
            print(name)
            current_row = updated_names.index(name) + 2
            current_column = 2
            (sheet.cell(row=current_row, column=current_column)).value = name
            for presence in updated_record[name]:
                current_column += 1
                (sheet.cell(row=current_row, column=current_column)).value = presence

        wb.save(excel_file)
        FileManager.start_file(excel_file)

    @staticmethod
    def process_names(names):
        names = names.strip().split(',')
        for index, name in enumerate(names):
            names[index] = name.strip()
            if name == '':
                names.remove(name)
        names = list(set(names))  # remove repeating names

        # sorting names using insertion sort
        for i in range(1, len(names)):
            key = names[i]
            j = i - 1

            while j >= 0 and key < names[j]:
                names[j + 1] = names[j]
                j -= 1

            names[j + 1] = key

        return names

    @staticmethod
    def return_record_names(records):
        record_names_tmp = []
        record_names = []
        for record in records:
            record_names_tmp.append(ntpath.basename(record))

        for record in record_names_tmp:
            record_name = record.replace('_ta_app.xlsx', '')
            record_names.append(record_name)

        return record_names

    @staticmethod
    def start_file(file_path):
        if sys.platform == "win32":
            os.startfile(file_path)
        else:
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, file_path])

    @staticmethod
    def save_recent_searches(search):
        if not os.path.exists(f'users/{User.current_user}/rsp.pickle'):
            search_data = open(f'users/{User.current_user}/rsp.pickle', 'wb')
            pickle.dump([search], search_data)
            search_data.close()
        else:
            new_searches = [search]
            with open(f'users/{User.current_user}/rsp.pickle', 'rb') as recent_searches:
                for elements in pickle.load(recent_searches):
                    new_searches.append(elements)

            with open(f'users/{User.current_user}/rsp.pickle', 'wb') as recent_searches:
                pickle.dump(new_searches, recent_searches)

    @staticmethod
    def get_recent_searches():
        if not os.path.exists(f'users/{User.current_user}/rsp.pickle'):
            empty = ['', '', '']
            return tuple(empty)

        searches = []

        with open(f'users/{User.current_user}/rsp.pickle', 'rb') as recent_searches:
            for elements in pickle.load(recent_searches):
                searches.append(elements)

        searches[:] = [search for search in searches if search != '']

        if len(searches) == 1:
            one_search = []
            for elements in searches:
                one_search.append(elements)

            one_search.append('')
            one_search.append('')
            return tuple(one_search)

        if len(searches) == 2:
            two_searches = []
            for elements in searches:
                two_searches.append(elements)
            two_searches.append('')
            return tuple(two_searches)

        if len(searches) > 3:
            return tuple(searches[:3])

        return tuple(searches)

    @staticmethod
    def view_record(instance):
        file_path = FileManager.find_file(instance.text)[0]
        FileManager.start_file(file_path)

    @staticmethod
    def delete_record(instance):
        file_to_delete = FileManager.find_file(instance.id)[0]
        os.remove(file_to_delete)


class AttendanceHandler:
    def __init__(self, excel_sheet):
        self.excel_names = []
        self.excel_sheet = excel_sheet

    def get_present_students(self, chat_file):
        self.get_names_excel()
        present_students_tmp = []
        with open(chat_file) as chat_file:
            for line in chat_file:
                present_students_tmp.append([student for student in self.excel_names if f'From {student}' in line])
        present_students = [students for lists in present_students_tmp for students in lists]
        return present_students

    def get_names_excel(self):
        max_row = self.excel_sheet.max_row
        current_row = 2
        while current_row <= max_row:
            self.excel_names.append((self.excel_sheet.cell(row=current_row, column=2)).value)
            current_row += 1

